from screener import report


def test_score_color_none_is_grey():
    assert report._score_color(None) == "#eeeeee"


def test_score_color_returns_hex():
    c = report._score_color(50, 100)
    assert c.startswith("#") and len(c) == 7


def test_score_color_high_is_greener_than_low():
    low = report._score_color(0, 100)
    high = report._score_color(100, 100)
    # 緑チャンネル(3-4文字目)が高スコアで大きい
    assert int(high[3:5], 16) > int(low[3:5], 16)


def test_svg_hbar_contains_labels():
    svg = report._svg_hbar([("デンソー", 86.0), ("トヨタ", 50.0)], 100)
    assert svg.startswith("<svg") and svg.endswith("</svg>")
    assert "デンソー" in svg and "トヨタ" in svg


def test_svg_hbar_empty_is_blank():
    assert report._svg_hbar([], 100) == ""


def test_svg_gauge_shows_score():
    svg = report._svg_gauge(88)
    assert svg.startswith("<svg") and "88" in svg


def test_svg_gauge_none_is_zero():
    svg = report._svg_gauge(None)
    assert ">0<" in svg


def test_build_html_writes_file(tmp_path, sections, market):
    out = report.build_html(sections, market, tmp_path / "r.html")
    assert out.exists()
    text = out.read_text(encoding="utf-8")
    assert "DENSO CORP" in text          # バリュー銘柄
    assert "NTT INC" in text             # 逆張り銘柄
    assert "バリュー" in text and "逆張り" in text and "モメンタム" in text
    assert "極度の強欲" in text          # 市況ラベル


def test_build_html_empty_section_shows_none(tmp_path, sections, market):
    out = report.build_html(sections, market, tmp_path / "r.html")
    # momentum は空 → 「該当なし」
    assert "該当なし" in out.read_text(encoding="utf-8")


def test_build_excel_sheets(tmp_path, sections, market):
    import openpyxl
    out = report.build_excel(sections, market, tmp_path / "r.xlsx")
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    for s in ["サマリ", "バリュー", "逆張り", "モメンタム", "市況"]:
        assert s in wb.sheetnames


def test_build_excel_value_has_ticker(tmp_path, sections, market):
    import openpyxl
    out = report.build_excel(sections, market, tmp_path / "r.xlsx")
    wb = openpyxl.load_workbook(out)
    vals = [c.value for row in wb["バリュー"].iter_rows() for c in row]
    assert "6902.T" in vals and "score（点）" in vals  # ヘッダは単位付き


def test_build_excel_empty_sheet(tmp_path, sections, market):
    import openpyxl
    out = report.build_excel(sections, market, tmp_path / "r.xlsx")
    wb = openpyxl.load_workbook(out)
    assert wb["モメンタム"]["A1"].value == "該当なし"


def test_svg_line_draws_polyline():
    svg = report._svg_line([100, 110, 105, 120])
    assert svg.startswith("<svg") and "polyline" in svg


def test_svg_line_too_short_is_blank():
    assert report._svg_line([100]) == ""


def test_table_html_link_base_links_ticker():
    rows = [{"ticker": "7203.T", "name": "TOYOTA", "score": 80}]
    html_out = report._table_html(rows, 10, 100, link_base="/stock/")
    assert '<a href="/stock/7203.T">' in html_out


def test_table_html_no_link_by_default():
    rows = [{"ticker": "7203.T", "score": 80}]
    assert "<a href" not in report._table_html(rows, 10, 100)


def test_header_label_adds_units():
    assert report.header_label("score") == "score（点）"
    assert report.header_label("PER") == "PER（倍）"
    assert report.header_label("出来高比") == "出来高比（倍）"


def test_header_label_keeps_percent_columns():
    assert report.header_label("配当%") == "配当%"
    assert report.header_label("RSI") == "RSI"


def test_table_html_header_has_unit():
    rows = [{"ticker": "7203.T", "score": 80, "PER": 9.4}]
    out = report._table_html(rows, 10, 100)
    assert "score（点）" in out and "PER（倍）" in out


def test_render_html_returns_string(sections, market):
    out = report.render_html(sections, market, header_extra="<p>DL_LINK</p>")
    assert isinstance(out, str)
    assert out.startswith("<!doctype html>")
    assert "DENSO CORP" in out
    assert "DL_LINK" in out


def test_build_workbook_returns_wb(sections, market):
    wb = report._build_workbook(sections, market)
    assert "バリュー" in wb.sheetnames and "市況" in wb.sheetnames
