"""HTML / Excel レポート生成。

スクリーニング結果(セクション)と市況(Fear&Greed)を、色付き表 +
インラインSVGグラフの 1ファイル HTML と、書式付き Excel に出力する。
"""
from __future__ import annotations

import html
import math
from pathlib import Path


def _mix(c1, c2, t):
    """2色(0-1のRGBタプル)を t で線形補間。"""
    return tuple(a + (b - a) * t for a, b in zip(c1, c2))


def _score_color(score, max_score=100):
    """スコアを赤(低)→黄→緑(高)の hex 色 "#rrggbb" に変換。"""
    if score is None:
        return "#eeeeee"
    f = score / max_score if max_score else 0
    f = max(0.0, min(1.0, f))
    red, yellow, green = (0.85, 0.19, 0.15), (1.0, 0.85, 0.20), (0.10, 0.60, 0.31)
    if f < 0.5:
        r, g, b = _mix(red, yellow, f / 0.5)
    else:
        r, g, b = _mix(yellow, green, (f - 0.5) / 0.5)
    return f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}"


def _svg_hbar(items, max_val, width=380, bar_h=22, gap=6):
    """items=[(label, value)] の横棒グラフ SVG を返す。空なら ""。"""
    if not items:
        return ""
    n = len(items)
    height = n * (bar_h + gap) + gap
    label_w = 130
    bar_area = width - label_w - 45
    out = []
    for i, (label, value) in enumerate(items):
        y = gap + i * (bar_h + gap)
        v = value if (value is not None and value == value) else 0
        w = max(2, bar_area * (v / max_val if max_val else 0))
        color = _score_color(v, max_val)
        lab = html.escape(str(label))
        out.append(
            f'<text x="0" y="{y+bar_h*0.7:.0f}" font-size="12">{lab}</text>'
            f'<rect x="{label_w}" y="{y}" width="{w:.0f}" height="{bar_h}" '
            f'fill="{color}" rx="3"/>'
            f'<text x="{label_w+w+5:.0f}" y="{y+bar_h*0.7:.0f}" '
            f'font-size="11">{v}</text>'
        )
    return (f'<svg viewBox="0 0 {width} {height}" width="{width}" '
            f'xmlns="http://www.w3.org/2000/svg">' + "".join(out) + "</svg>")


def _svg_gauge(score):
    """0-100 の半円ゲージ SVG。score=None は 0 扱い。"""
    s = 0 if score is None else max(0, min(100, score))
    angle = math.pi * (1 - s / 100)  # 0点=左(180°), 100点=右(0°)
    cx, cy, r = 110, 110, 90
    x = cx + r * math.cos(angle)
    y = cy - r * math.sin(angle)
    return (
        '<svg viewBox="0 0 220 130" width="220" '
        'xmlns="http://www.w3.org/2000/svg">'
        '<path d="M20 110 A90 90 0 0 1 200 110" fill="none" '
        'stroke="#ddd" stroke-width="16"/>'
        f'<circle cx="{x:.1f}" cy="{y:.1f}" r="9" fill="{_score_color(s)}"/>'
        f'<text x="110" y="100" font-size="30" text-anchor="middle" '
        f'font-weight="bold">{int(s)}</text>'
        '</svg>'
    )


def _svg_line(series, width=640, height=180, pad=24):
    """数値系列(時系列)の終値折れ線 SVG。点が2未満なら ""。"""
    vals = [float(v) for v in list(series) if v is not None and v == v]
    if len(vals) < 2:
        return ""
    lo, hi = min(vals), max(vals)
    rng = (hi - lo) or 1
    n = len(vals)
    pts = " ".join(
        f"{pad + (width - 2 * pad) * i / (n - 1):.1f},"
        f"{height - pad - (height - 2 * pad) * (v - lo) / rng:.1f}"
        for i, v in enumerate(vals)
    )
    return (
        f'<svg viewBox="0 0 {width} {height}" width="{width}" '
        'xmlns="http://www.w3.org/2000/svg">'
        f'<polyline points="{pts}" fill="none" stroke="#305496" stroke-width="2"/>'
        f'<text x="{pad}" y="14" font-size="11">高 {hi:.0f}</text>'
        f'<text x="{pad}" y="{height - 6}" font-size="11">安 {lo:.0f}</text>'
        '</svg>'
    )


SECTION_META = {
    "value":      ("バリュースクリーニング（割安株 / 100点）", 100, "バリュー"),
    "contrarian": ("逆張り（売られすぎ / 0-6）", 6, "逆張り"),
    "momentum":   ("モメンタム（強い銘柄 / 0-5）", 5, "モメンタム"),
}

_HTML_HEAD = """<!doctype html><html lang="ja"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>株スクリーニングレポート</title>
<style>
 body{font-family:'Segoe UI','Meiryo',sans-serif;margin:24px;color:#222;background:#fafafa}
 h1{font-size:22px} h2{font-size:17px;border-left:5px solid #305496;padding-left:8px;margin-top:28px}
 section{background:#fff;border:1px solid #e3e3e3;border-radius:8px;padding:14px 18px;margin:14px 0}
 table{border-collapse:collapse;width:100%;font-size:13px;margin:8px 0}
 th{background:#305496;color:#fff;padding:6px 8px;text-align:left;white-space:nowrap}
 td{border-bottom:1px solid #eee;padding:5px 8px}
 .empty{color:#999;font-style:italic}
 .gauge{text-align:center} .gauge .label{font-weight:bold;font-size:15px;margin:4px}
</style></head><body>"""


UNITS = {"score": "点", "value": "点", "change": "点",
         "PER": "倍", "PBR": "倍", "出来高比": "倍"}


def header_label(col):
    """表示用の列見出し。単位があれば付ける（内部キーは不変）。"""
    u = UNITS.get(col)
    return f"{col}（{u}）" if u else str(col)


def _fmt(v):
    if isinstance(v, list):
        return " / ".join(v)
    return "" if v is None else str(v)


def _table_html(rows, top, max_val, link_base=None):
    if not rows:
        return "<p class='empty'>該当なし</p>"
    rows = rows[:top]
    cols = list(rows[0].keys())
    head = "".join(f"<th>{html.escape(header_label(c))}</th>" for c in cols)
    body = []
    for r in rows:
        tds = []
        for c in cols:
            v = r.get(c)
            cell = html.escape(_fmt(v))
            if c == "score":
                tds.append(f'<td style="background:{_score_color(v, max_val)};'
                           f'text-align:right;font-weight:bold">{cell}</td>')
            elif c == "ticker" and link_base:
                href = html.escape(link_base) + cell
                tds.append(f'<td><a href="{href}">{cell}</a></td>')
            else:
                tds.append(f"<td>{cell}</td>")
        body.append("<tr>" + "".join(tds) + "</tr>")
    return (f"<table><thead><tr>{head}</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>")


def build_html(sections, market, path, top=20):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    p = [_HTML_HEAD, "<h1>株スクリーニングレポート</h1>"]

    # 市況
    p.append("<section><h2>市場センチメント（Fear &amp; Greed）</h2>")
    if market.get("score") is not None:
        vix = market.get("VIX")
        vix_txt = f"（VIX {html.escape(str(vix))}）" if vix is not None else ""
        p.append(f"<div class='gauge'>{_svg_gauge(market['score'])}"
                 f"<p class='label'>{html.escape(str(market.get('label','')))}"
                 f"{vix_txt}</p></div>")
        uw = market.get("内訳", {})
        p.append(_svg_hbar([(k, v) for k, v in uw.items()], 100))
    else:
        p.append("<p class='empty'>市況データ取得失敗</p>")
    p.append("</section>")

    # 各スクリーニング
    for key, (title, mv, _sheet) in SECTION_META.items():
        rows = sections.get(key, [])
        p.append(f"<section><h2>{html.escape(title)}</h2>")
        p.append(_table_html(rows, top, mv))
        bar = [(r.get("name") or r["ticker"], r["score"]) for r in rows[:10]]
        p.append(_svg_hbar(bar, mv))
        p.append("</section>")

    p.append("</body></html>")
    path.write_text("\n".join(p), encoding="utf-8")
    return path


def build_excel(sections, market, path, top=20):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # サマリシート
    ws = wb.active
    ws.title = "サマリ"
    ws["A1"] = "株スクリーニングレポート"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "市況 Fear&Greed"; ws["B3"] = market.get("score")
    ws["A4"] = "判定";            ws["B4"] = market.get("label")
    ws["A5"] = "VIX";             ws["B5"] = market.get("VIX")
    row = 7
    for key, (title, mv, _sheet) in SECTION_META.items():
        ws.cell(row, 1, title).font = Font(bold=True)
        ws.cell(row, 2, f"{len(sections.get(key, []))} 件")
        row += 1
    ws.column_dimensions["A"].width = 36

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")

    # 各スクリーニングシート
    for key, (title, mv, sheet) in SECTION_META.items():
        rows = sections.get(key, [])[:top]
        ws = wb.create_sheet(sheet)
        if not rows:
            ws["A1"] = "該当なし"
            continue
        cols = list(rows[0].keys())
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(1, ci, c)
            cell.fill = header_fill
            cell.font = header_font
        for ri, r in enumerate(rows, 2):
            for ci, c in enumerate(cols, 1):
                v = r.get(c)
                ws.cell(ri, ci, " / ".join(v) if isinstance(v, list) else v)
        ws.freeze_panes = "A2"
        # スコア列カラースケール
        if "score" in cols:
            col = get_column_letter(cols.index("score") + 1)
            rng = f"{col}2:{col}{len(rows) + 1}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=mv / 2, mid_color="FFEB84",
                end_type="num", end_value=mv, end_color="63BE7B"))

    # 市況シート（内訳）
    ws = wb.create_sheet("市況")
    ws["A1"] = "指標"; ws["B1"] = "スコア"
    ws["A1"].fill = ws["B1"].fill = header_fill
    ws["A1"].font = ws["B1"].font = header_font
    for i, (k, v) in enumerate(market.get("内訳", {}).items(), 2):
        ws.cell(i, 1, k); ws.cell(i, 2, v)

    # バリューに棒グラフ（スコア上位）
    vrows = sections.get("value", [])[:top]
    if vrows and "score" in vrows[0]:
        vs = wb["バリュー"]
        cols = list(vrows[0].keys())
        sc = cols.index("score") + 1
        namec = (cols.index("name") + 1) if "name" in cols else 1
        chart = BarChart()
        chart.type = "bar"
        chart.title = "バリュースコア"
        data = Reference(vs, min_col=sc, min_row=1, max_row=len(vrows) + 1)
        cats = Reference(vs, min_col=namec, min_row=2, max_row=len(vrows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8; chart.width = 16
        vs.add_chart(chart, f"{get_column_letter(len(cols) + 2)}2")

    wb.save(path)
    return path
